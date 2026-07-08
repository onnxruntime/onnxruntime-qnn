#!/usr/bin/env python3
# Copyright (c) Qualcomm Technologies, Inc. and/or its subsidiaries.
# SPDX-License-Identifier: MIT
#
# Generate a per-day coverage snapshot Excel from a Cobertura XML report.
# Each nightly run creates an independent file.
# TODO: enable the aggregate CI workflow to merge daily snapshots into a 30-day history report.
#
# Usage:
#   python3 update_coverage_excel.py \
#       --coverage-xml <path/to/coverage.xml> \
#       --excel        <path/to/per_file_coverage.xlsx> \
#       --commit-sha   <40-char SHA> \
#       [--date        YYYY-MM-DD]         # default: today (UTC)

import argparse
import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl is not installed. Activate the project venv or: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SHEET_NAME = "History"
HEADER = ["date", "commit_sha", "file", "line_pct", "branch_pct"]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Generate a per-day per-file coverage snapshot Excel from a Cobertura XML report."
    )
    p.add_argument("--coverage-xml", required=True, type=Path, help="Path to Cobertura XML (coverage.xml)")
    p.add_argument("--excel", required=True, type=Path, help="Output Excel path (overwritten if exists)")
    p.add_argument("--commit-sha", required=True, help="Current commit SHA (full 40-char)")
    p.add_argument(
        "--date",
        default=None,
        help="Override date as YYYY-MM-DD (default: $COVERAGE_RUN_DATE if set, else today UTC)",
    )
    return p.parse_args()


def parse_coverage_xml(xml_path: Path) -> list[dict]:
    if not xml_path.is_file():
        print(f"ERROR: coverage XML not found: {xml_path}", file=sys.stderr)
        sys.exit(1)

    try:
        tree = ET.parse(xml_path)
    except ET.ParseError as e:
        print(f"ERROR: failed to parse coverage XML {xml_path}: {e}", file=sys.stderr)
        sys.exit(1)
    root = tree.getroot()

    rows: list[dict] = []
    for cls in root.iter("class"):
        filename = cls.get("filename", "").strip()
        if not filename:
            continue
        rows.append(
            {
                "file": filename,
                "line_pct": round(float(cls.get("line-rate", "0")) * 100, 1),
                "branch_pct": round(float(cls.get("branch-rate", "0")) * 100, 1),
            }
        )

    rows.sort(key=lambda r: r["file"])
    return rows


def create_workbook(run_date: str, commit_sha: str, file_rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(HEADER)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in file_rows:
        ws.append([run_date, commit_sha, r["file"], r["line_pct"], r["branch_pct"]])

    return wb


def main() -> None:
    args = parse_args()

    # Date source-of-truth order: CLI --date > $COVERAGE_RUN_DATE env > today UTC.
    # The env var lets CI pass a date computed once upstream so multiple consumers
    # (this script + downstream upload path) agree even across UTC midnight.
    run_date = args.date or os.environ.get("COVERAGE_RUN_DATE") or datetime.now(timezone.utc).strftime("%Y-%m-%d")

    print(f"Coverage XML : {args.coverage_xml}")
    print(f"Excel        : {args.excel}")
    print(f"Commit SHA   : {args.commit_sha}")
    print(f"Date         : {run_date}")

    file_rows = parse_coverage_xml(args.coverage_xml)
    if not file_rows:
        print("WARNING: coverage XML contains zero <class> elements — Excel will be empty", file=sys.stderr)
    print(f"Files parsed : {len(file_rows)}")

    args.excel.parent.mkdir(parents=True, exist_ok=True)
    wb = create_workbook(run_date, args.commit_sha, file_rows)
    wb.save(args.excel)
    print(f"Saved        : {args.excel}")


if __name__ == "__main__":
    main()
